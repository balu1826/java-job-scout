"""
Apify Actor — Java Job Scout (Full Pipeline)
============================================
This file runs INSIDE Apify's cloud. It:
  1. Scrapes LinkedIn jobs via Apify API
  2. Analyzes them with Groq (free)
  3. Generates Excel report
  4. Emails it to you

Deploy this as a custom Apify Actor and schedule it at 10 AM IST daily.
All secrets are stored in Apify's input schema (encrypted).
"""

import os, sys, json, time, smtplib, requests
from datetime import date
from pathlib import Path
from email.mime.multipart import MIMEMultipart
from email.mime.text      import MIMEText
from email.mime.base      import MIMEBase
from email                import encoders
from openpyxl             import Workbook
from openpyxl.styles      import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils       import get_column_letter

# ── Apify provides input via environment or Actor input ───────────────────────
try:
    from apify_client import ApifyClient
    INPUT = ApifyClient(os.environ.get("APIFY_TOKEN","")).get_input() or {}
except Exception:
    INPUT = {}

APIFY_TOKEN      = INPUT.get("apifyToken")      or os.environ.get("APIFY_API_TOKEN","")
GROQ_API_KEY     = INPUT.get("groqApiKey")      or os.environ.get("GROQ_API_KEY","")
GMAIL_SENDER     = INPUT.get("gmailSender")     or os.environ.get("GMAIL_SENDER","")
GMAIL_PASSWORD   = INPUT.get("gmailPassword")   or os.environ.get("GMAIL_APP_PASSWORD","")
RECIPIENT_EMAIL  = INPUT.get("recipientEmail")  or os.environ.get("RECIPIENT_EMAIL","")

GROQ_MODEL  = "llama-3.3-70b-versatile"
GROQ_URL    = "https://api.groq.com/openai/v1/chat/completions"
ACTOR_ID    = "curious_coder/linkedin-jobs-scraper"
BASE_URL    = "https://api.apify.com/v2"

SEARCH_URLS = [
    "https://www.linkedin.com/jobs/search/?keywords=Java+Developer+Fresher&location=India&f_TPR=r86400&f_E=1&sortBy=DD&position=1&pageNum=0",
    "https://www.linkedin.com/jobs/search/?keywords=Associate+Software+Developer+Java+Spring+Boot&location=India&f_TPR=r86400&f_E=1&sortBy=DD&position=1&pageNum=0",
    "https://www.linkedin.com/jobs/search/?keywords=Software+Engineer+Java+Backend+Fresher&location=India&f_TPR=r86400&f_E=1&sortBy=DD&position=1&pageNum=0",
]

RESUME_TEXT = """
Battula Bala Kotaiah — Java Backend Developer (Fresher)
Stack: Java, Spring Boot, Spring Data JPA, MySQL, PostgreSQL, JWT, RBAC/PBAC, REST APIs
Experience: 6 months apprenticeship at Tekworks (Dec 2025 – May 2026)
  - Built RBAC/PBAC authorization framework
  - Audit logging for entity-level changes
  - Custom interceptors for rate limiting and API monitoring
Projects:
  - Doctor Appointment System (Spring Boot, JPA, MySQL, JWT, CRUD REST APIs)
  - ATS Resume Builder (Spring Boot, Groq AI, REST APIs, HTML/CSS generation)
Education: BTech CS, RK College of Engineering (CGPA 7.47)
"""

# ══════════════════════════════════════════════════════════════════════════════
# STEP 1 — SCRAPE
# ══════════════════════════════════════════════════════════════════════════════
def scrape_jobs() -> list[dict]:
    print("📡 Scraping LinkedIn via Apify ...")
    run = requests.post(
        f"{BASE_URL}/acts/{ACTOR_ID}/runs",
        params={"token": APIFY_TOKEN},
        json={"urls": SEARCH_URLS, "count": 30, "scrapeCompany": False},
        timeout=30,
    )
    run.raise_for_status()
    run_id = run.json()["data"]["id"]

    for i in range(30):
        time.sleep(10)
        s = requests.get(f"{BASE_URL}/actor-runs/{run_id}", params={"token": APIFY_TOKEN}, timeout=15)
        status = s.json()["data"]["status"]
        print(f"  attempt {i+1}: {status}")
        if status == "SUCCEEDED":
            break
        if status in ("FAILED","ABORTED","TIMED-OUT"):
            raise RuntimeError(f"Apify run {status}")

    ds_id = s.json()["data"]["defaultDatasetId"]
    items = requests.get(
        f"{BASE_URL}/datasets/{ds_id}/items",
        params={"token": APIFY_TOKEN, "limit": 30,
                "fields": "title,companyName,location,link,postedAt,seniorityLevel,employmentType,descriptionText,applicantsCount"},
        timeout=30,
    )
    items.raise_for_status()
    jobs = items.json()
    print(f"  → {len(jobs)} jobs fetched")
    return jobs


# ══════════════════════════════════════════════════════════════════════════════
# STEP 2 — ANALYZE with Groq
# ══════════════════════════════════════════════════════════════════════════════
def analyze_jobs(jobs: list[dict]) -> list[dict]:
    print("🤖 Analyzing with Groq llama-3.3-70b ...")

    slim = [{
        "title":           j.get("title",""),
        "company":         j.get("companyName",""),
        "location":        j.get("location",""),
        "link":            j.get("link",""),
        "date_posted":     str(j.get("postedAt","")),
        "seniority":       j.get("seniorityLevel",""),
        "employment_type": j.get("employmentType",""),
        "description":     (j.get("descriptionText") or "")[:800],
    } for j in jobs]

    prompt = f"""You are a brutally honest career advisor for Java developer roles in India.
Candidate resume:
{RESUME_TEXT}

Analyze these {len(slim)} jobs and return ONLY a JSON array, no markdown:
{json.dumps(slim, indent=2)}

For each job return:
- title, company, location, link, date_posted, seniority, employment_type (strings)
- stars: 1-5 integer match score
- star_label: "★★★★★" style string
- skills_matched: list of skills candidate HAS
- skills_missing: list of skills candidate is MISSING
- is_fresher_friendly: boolean
- fresher_reason: one sentence
- verdict: one brutally honest sentence
- apply_priority: "APPLY NOW" | "STRONG CONSIDER" | "STRETCH GOAL" | "SKIP"

Sort by stars descending. Return ONLY the JSON array."""

    r = requests.post(
        GROQ_URL,
        headers={"Authorization": f"Bearer {GROQ_API_KEY}", "Content-Type": "application/json"},
        json={"model": GROQ_MODEL, "temperature": 0.3, "max_tokens": 8000,
              "messages": [{"role": "user", "content": prompt}]},
        timeout=120,
    )
    r.raise_for_status()
    raw = r.json()["choices"][0]["message"]["content"].strip()
    if raw.startswith("```"):
        raw = raw.split("\n",1)[1].rsplit("```",1)[0]
    analyzed = json.loads(raw)
    print(f"  → {len(analyzed)} jobs analyzed")
    return analyzed


# ══════════════════════════════════════════════════════════════════════════════
# STEP 3 — GENERATE EXCEL
# ══════════════════════════════════════════════════════════════════════════════
BORDER = Border(left=Side(style="thin"), right=Side(style="thin"),
                top=Side(style="thin"),  bottom=Side(style="thin"))

STAR_FILLS = {
    5: PatternFill("solid", fgColor="C6EFCE"),
    4: PatternFill("solid", fgColor="DDEBF7"),
    3: PatternFill("solid", fgColor="FFEB9C"),
    2: PatternFill("solid", fgColor="FFCCCC"),
    1: PatternFill("solid", fgColor="FF9999"),
}

def _hdr(ws, r, c, v, bg="1F3864", fg="FFFFFF"):
    cell = ws.cell(row=r, column=c, value=v)
    cell.font      = Font(bold=True, color=fg, name="Arial", size=10)
    cell.fill      = PatternFill("solid", fgColor=bg)
    cell.border    = BORDER
    cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

def _dat(ws, r, c, v, fill=None, align="left"):
    cell = ws.cell(row=r, column=c, value=v)
    cell.font      = Font(name="Arial", size=9)
    cell.border    = BORDER
    cell.alignment = Alignment(wrap_text=True, vertical="top", horizontal=align)
    if fill: cell.fill = fill

def _lnk(ws, r, c, url, fill=None):
    cell = ws.cell(row=r, column=c, value="🔗 Open")
    cell.hyperlink  = url
    cell.font       = Font(name="Arial", size=9, color="0563C1", underline="single")
    cell.border     = BORDER
    cell.alignment  = Alignment(horizontal="center", vertical="top")
    if fill: cell.fill = fill

def _widths(ws, widths):
    for i,w in enumerate(widths,1):
        ws.column_dimensions[get_column_letter(i)].width = w

def generate_excel(jobs: list[dict]) -> str:
    today    = date.today().strftime("%B %d, %Y")
    out_path = f"daily_jobs_{date.today().strftime('%Y-%m-%d')}.xlsx"
    wb       = Workbook()

    # Sheet 1 — Top Matches
    ws1 = wb.active
    ws1.title = "Top Job Matches"
    ws1.merge_cells("A1:J1")
    ws1["A1"].value     = f"🎯 Daily Java Job Matches — Battula Bala Kotaiah | {today}"
    ws1["A1"].font      = Font(bold=True, color="1F3864", size=13, name="Arial")
    ws1["A1"].alignment = Alignment(horizontal="center", vertical="center")
    ws1.row_dimensions[1].height = 28

    ws1.merge_cells("A2:J2")
    ws1["A2"].value     = "Groq llama-3.3-70b analyzed every JD against your resume  |  Stack: Java · Spring Boot · JPA · JWT · MySQL · PostgreSQL"
    ws1["A2"].font      = Font(italic=True, color="555555", size=9, name="Arial")
    ws1["A2"].alignment = Alignment(horizontal="center")

    h1 = ["#","Job Title","Company","Location","Date Posted","Level","Match","Apply Priority","Missing Skills","Link"]
    w1 = [4, 42, 28, 26, 14, 16, 10, 18, 38, 12]
    _widths(ws1, w1)
    for ci,h in enumerate(h1,1): _hdr(ws1, 3, ci, h)
    ws1.row_dimensions[3].height = 22

    for ri, job in enumerate(jobs, start=4):
        sf      = STAR_FILLS.get(job.get("stars",0), PatternFill())
        missing = ", ".join(job.get("skills_missing",[])) or "—"
        _dat(ws1, ri, 1,  ri-3,                         fill=sf, align="center")
        _dat(ws1, ri, 2,  job.get("title",""),           fill=sf)
        _dat(ws1, ri, 3,  job.get("company",""),         fill=sf)
        _dat(ws1, ri, 4,  job.get("location",""),        fill=sf)
        _dat(ws1, ri, 5,  job.get("date_posted",""),     fill=sf, align="center")
        _dat(ws1, ri, 6,  job.get("seniority",""),       fill=sf)
        _dat(ws1, ri, 7,  job.get("star_label",""),      fill=sf, align="center")
        _dat(ws1, ri, 8,  job.get("apply_priority",""),  fill=sf, align="center")
        _dat(ws1, ri, 9,  missing,                       fill=sf)
        _lnk(ws1, ri, 10, job.get("link",""),            fill=sf)
        ws1.row_dimensions[ri].height = 32
    ws1.freeze_panes = "A4"

    # Sheet 2 — Fresher Only
    ws2      = wb.create_sheet("Fresher Only 🟢")
    freshers = [j for j in jobs if j.get("is_fresher_friendly")]
    ws2.merge_cells("A1:H1")
    ws2["A1"].value     = f"🟢 Fresher & Entry-Level Only — {len(freshers)} jobs today"
    ws2["A1"].font      = Font(bold=True, color="375623", size=12, name="Arial")
    ws2["A1"].alignment = Alignment(horizontal="center", vertical="center")
    ws2.row_dimensions[1].height = 26

    h2 = ["#","Job Title","Company","Location","Date Posted","Level","Why Fresher-Friendly","Link"]
    w2 = [4, 42, 28, 26, 14, 16, 42, 12]
    _widths(ws2, w2)
    for ci,h in enumerate(h2,1): _hdr(ws2, 2, ci, h, bg="375623")

    alt = [PatternFill("solid", fgColor="E2EFDA"), PatternFill("solid", fgColor="FFFFFF")]
    for ri, job in enumerate(freshers, start=3):
        sf = alt[ri % 2]
        _dat(ws2, ri, 1, ri-2,                        fill=sf, align="center")
        _dat(ws2, ri, 2, job.get("title",""),          fill=sf)
        _dat(ws2, ri, 3, job.get("company",""),        fill=sf)
        _dat(ws2, ri, 4, job.get("location",""),       fill=sf)
        _dat(ws2, ri, 5, job.get("date_posted",""),    fill=sf, align="center")
        _dat(ws2, ri, 6, job.get("seniority",""),      fill=sf)
        _dat(ws2, ri, 7, job.get("fresher_reason",""), fill=sf)
        _lnk(ws2, ri, 8, job.get("link",""),           fill=sf)
        ws2.row_dimensions[ri].height = 30
    ws2.freeze_panes = "A3"

    # Sheet 3 — Skills Gap
    from collections import Counter
    ws3     = wb.create_sheet("Skills Gap 🔴")
    counter = Counter()
    for job in jobs:
        for s in job.get("skills_missing",[]): counter[s.strip()] += 1
    total   = len(jobs)

    ws3.merge_cells("A1:D1")
    ws3["A1"].value     = f"🔴 Aggregated Skills Gap — {total} jobs · {today}"
    ws3["A1"].font      = Font(bold=True, color="C00000", size=12, name="Arial")
    ws3["A1"].alignment = Alignment(horizontal="center", vertical="center")

    h3 = ["Missing Skill", f"Seen in X of {total} Jobs", "Frequency %", "Priority"]
    w3 = [38, 20, 14, 45]
    _widths(ws3, w3)
    for ci,h in enumerate(h3,1): _hdr(ws3, 2, ci, h, bg="C00000")

    gap_fills = {
        70: PatternFill("solid", fgColor="FCE4D6"),
        50: PatternFill("solid", fgColor="FFF2CC"),
        30: PatternFill("solid", fgColor="FFFACD"),
         0: PatternFill("solid", fgColor="E2EFDA"),
    }
    for ri, (skill, count) in enumerate(counter.most_common(), start=3):
        pct  = round(count / total * 100)
        prio = ("🔴 Critical — Learn ASAP"           if pct >= 70 else
                "🟠 High — Learn within 1 month"      if pct >= 50 else
                "🟡 Medium — Learn within 3 months"   if pct >= 30 else
                "🟢 Nice to have")
        sf   = next(v for k,v in sorted(gap_fills.items(), reverse=True) if pct >= k)
        _dat(ws3, ri, 1, skill,             fill=sf)
        _dat(ws3, ri, 2, f"{count}/{total}", fill=sf, align="center")
        _dat(ws3, ri, 3, f"{pct}%",         fill=sf, align="center")
        _dat(ws3, ri, 4, prio,              fill=sf)
        ws3.row_dimensions[ri].height = 22
    ws3.freeze_panes = "A3"

    # Sheet 4 — Summary
    ws4 = wb.create_sheet("Summary & Plan")
    ws4.column_dimensions["A"].width = 3
    ws4.column_dimensions["B"].width = 65
    ws4.merge_cells("A1:B1")
    ws4["A1"].value     = f"📋 Daily Summary — {today}"
    ws4["A1"].font      = Font(bold=True, color="1F3864", size=13, name="Arial")
    ws4["A1"].alignment = Alignment(horizontal="center", vertical="center")
    ws4.row_dimensions[1].height = 30

    apply_now   = [j for j in jobs if j.get("apply_priority")=="APPLY NOW"]
    fresher_cnt = len(freshers)
    top5        = [j for j in jobs if j.get("stars",0)>=4][:5]

    lines = [
        ("TODAY'S SNAPSHOT", None, "1F3864", True),
        ("", f"Total jobs analyzed by Groq AI: {total}", None, False),
        ("", f"Fresher-friendly jobs: {fresher_cnt}", None, False),
        ("", f"APPLY NOW recommendations: {len(apply_now)}", None, False),
        ("", "", None, False),
        ("🏆 TOP 5 MATCHES TODAY", None, "375623", True),
    ]
    for i,j in enumerate(top5,1):
        lines.append(("", f"{i}. {j.get('title','')} — {j.get('company','')} | {j.get('star_label','')} | {j.get('apply_priority','')}", None, False))
    lines += [
        ("", "", None, False),
        ("✅ DO THIS TODAY", None, "375623", True),
        ("", "1️⃣  Apply to every APPLY NOW job before 12 PM.", None, False),
        ("", "2️⃣  Check Skills Gap sheet — focus on 🔴 Critical skills this week.", None, False),
        ("", "3️⃣  Push any new projects to GitHub before applying.", None, False),
        ("", "", None, False),
        ("📧 AUTOMATION INFO", None, "1F3864", True),
        ("", "Auto-generated by Groq llama-3.3-70b + Apify LinkedIn Scraper.", None, False),
        ("", "Runs every day at 10:00 AM IST. Jobs from last 24 hours only.", None, False),
        ("", "Total cost: ₹0 (Groq free tier + Apify free tier + GitHub free tier).", None, False),
    ]

    sec_fills = {
        "1F3864": PatternFill("solid", fgColor="D9E1F2"),
        "C00000": PatternFill("solid", fgColor="FCE4D6"),
        "375623": PatternFill("solid", fgColor="E2EFDA"),
    }
    row = 2
    for label, text, color, is_header in lines:
        if is_header:
            ws4.merge_cells(f"A{row}:B{row}")
            c = ws4.cell(row=row, column=1, value=f"  {label}")
            c.font      = Font(bold=True, color=color, size=11, name="Arial")
            c.fill      = sec_fills.get(color, PatternFill())
            c.alignment = Alignment(vertical="center")
            ws4.row_dimensions[row].height = 22
        else:
            ws4.cell(row=row, column=1, value="")
            c = ws4.cell(row=row, column=2, value=text)
            c.font      = Font(name="Arial", size=10)
            c.alignment = Alignment(wrap_text=True, vertical="top")
            ws4.row_dimensions[row].height = 18 if text else 8
        row += 1

    wb.save(out_path)
    print(f"  → Excel saved: {out_path}")
    return out_path


# ══════════════════════════════════════════════════════════════════════════════
# STEP 4 — SEND EMAIL
# ══════════════════════════════════════════════════════════════════════════════
def send_email(excel_path: str, jobs: list[dict]) -> None:
    today_str   = date.today().strftime("%B %d, %Y")
    apply_now   = sum(1 for j in jobs if j.get("apply_priority")=="APPLY NOW")
    fresher_cnt = sum(1 for j in jobs if j.get("is_fresher_friendly"))
    top5        = [j for j in jobs if j.get("stars",0)>=4][:5]

    rows = ""
    for i,j in enumerate(top5,1):
        bg = "#f0fff4" if i%2==0 else "#ffffff"
        pc = "#1a7a1a" if j.get("apply_priority")=="APPLY NOW" else "#b35900"
        rows += f"""<tr style="background:{bg}">
          <td style="padding:8px;border:1px solid #ddd">{i}</td>
          <td style="padding:8px;border:1px solid #ddd"><b>{j.get('title','')}</b></td>
          <td style="padding:8px;border:1px solid #ddd">{j.get('company','')}</td>
          <td style="padding:8px;border:1px solid #ddd">{j.get('location','')}</td>
          <td style="padding:8px;border:1px solid #ddd;text-align:center">{j.get('star_label','')}</td>
          <td style="padding:8px;border:1px solid #ddd;color:{pc};font-weight:bold;text-align:center">{j.get('apply_priority','')}</td>
          <td style="padding:8px;border:1px solid #ddd"><a href="{j.get('link','#')}" style="color:#0563C1">View</a></td>
        </tr>"""

    html = f"""<html><body style="font-family:Arial,sans-serif;color:#222;max-width:800px;margin:auto">
      <div style="background:#1F3864;color:#fff;padding:20px;border-radius:8px 8px 0 0">
        <h2 style="margin:0">🎯 Daily Java Job Report</h2>
        <p style="margin:4px 0;font-size:14px">{today_str} · Groq AI + Apify · Cost: ₹0</p>
      </div>
      <div style="background:#f8f9fa;padding:16px;border:1px solid #dee2e6">
        <b>Hi Balu 👋</b><br><br>
        Your daily Java job report is ready. Groq llama-3.3-70b analyzed every job against your resume.
        <br><br>
        <table style="border-collapse:collapse;width:100%">
          <tr>
            <td style="background:#E2EFDA;padding:12px;border-radius:6px;text-align:center;width:33%">
              <div style="font-size:24px;font-weight:bold;color:#375623">{len(jobs)}</div>
              <div style="font-size:12px;color:#555">Jobs Analyzed</div>
            </td>
            <td style="width:2%"></td>
            <td style="background:#DDEBF7;padding:12px;border-radius:6px;text-align:center;width:33%">
              <div style="font-size:24px;font-weight:bold;color:#1F3864">{fresher_cnt}</div>
              <div style="font-size:12px;color:#555">Fresher-Friendly</div>
            </td>
            <td style="width:2%"></td>
            <td style="background:#FCE4D6;padding:12px;border-radius:6px;text-align:center;width:30%">
              <div style="font-size:24px;font-weight:bold;color:#C00000">{apply_now}</div>
              <div style="font-size:12px;color:#555">Apply NOW</div>
            </td>
          </tr>
        </table>
      </div>
      <h3 style="color:#1F3864;margin:20px 0 8px">🏆 Top 5 Matches</h3>
      <table style="border-collapse:collapse;width:100%;font-size:13px">
        <tr style="background:#1F3864;color:#fff">
          <th style="padding:8px">#</th><th style="padding:8px">Role</th>
          <th style="padding:8px">Company</th><th style="padding:8px">Location</th>
          <th style="padding:8px">Match</th><th style="padding:8px">Priority</th>
          <th style="padding:8px">Link</th>
        </tr>{rows}
      </table>
      <div style="background:#fff8e1;border-left:4px solid #f59e0b;padding:12px;margin:20px 0;border-radius:0 6px 6px 0">
        <b>📎 Full Excel attached</b> — 4 sheets: Top Matches · Fresher Only · Skills Gap · Action Plan
      </div>
      <div style="background:#f1f5f9;padding:12px;border-radius:6px;font-size:11px;color:#666">
        Auto-generated · Groq llama-3.3-70b + Apify LinkedIn Scraper · Runs 10 AM IST daily · Cost: ₹0
      </div>
    </body></html>"""

    msg = MIMEMultipart("mixed")
    msg["Subject"] = f"🎯 Java Jobs {today_str} | {apply_now} APPLY NOW · {fresher_cnt} Fresher Friendly"
    msg["From"]    = GMAIL_SENDER
    msg["To"]      = RECIPIENT_EMAIL
    msg.attach(MIMEText(html, "html"))

    with open(excel_path, "rb") as f:
        part = MIMEBase("application", "octet-stream")
        part.set_payload(f.read())
    encoders.encode_base64(part)
    part.add_header("Content-Disposition", f'attachment; filename="{Path(excel_path).name}"')
    msg.attach(part)

    with smtplib.SMTP_SSL("smtp.gmail.com", 465) as smtp:
        smtp.login(GMAIL_SENDER, GMAIL_PASSWORD)
        smtp.sendmail(GMAIL_SENDER, RECIPIENT_EMAIL, msg.as_string())
    print(f"  → Email sent to {RECIPIENT_EMAIL} ✅")


# ══════════════════════════════════════════════════════════════════════════════
# MAIN
# ══════════════════════════════════════════════════════════════════════════════
if __name__ == "__main__":
    print("=" * 55)
    print(f"  Java Job Scout (Apify Actor) — {date.today()}")
    print(f"  AI: Groq {GROQ_MODEL} (FREE)")
    print("=" * 55)

    raw_jobs = scrape_jobs()
    if not raw_jobs:
        print("⚠️  No jobs found today.")
        sys.exit(0)

    analyzed   = analyze_jobs(raw_jobs)
    excel_path = generate_excel(analyzed)
    send_email(excel_path, analyzed)

    print("\n✅ All done! Check your inbox.")
    print("=" * 55)

"""
generate_excel.py
Builds a 4-sheet Excel report from Claude's analyzed job data.
Sheets:
  1. Top Job Matches   — all jobs sorted by stars
  2. Fresher Only 🟢   — only is_fresher_friendly=True
  3. Skills Gap 🔴     — aggregated missing skills across all jobs
  4. Summary & Plan    — action items
"""

from __future__ import annotations
from collections import Counter
from datetime import date
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter


# ── Styles ────────────────────────────────────────────────────────────────────
BORDER = Border(
    left=Side(style="thin"), right=Side(style="thin"),
    top=Side(style="thin"),  bottom=Side(style="thin"),
)

def hdr_cell(ws, row, col, value, bg="1F3864", fg="FFFFFF", size=10):
    c = ws.cell(row=row, column=col, value=value)
    c.font      = Font(bold=True, color=fg, name="Arial", size=size)
    c.fill      = PatternFill("solid", fgColor=bg)
    c.border    = BORDER
    c.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    return c

def data_cell(ws, row, col, value, fill=None, wrap=True, align="left", font_color="000000"):
    c = ws.cell(row=row, column=col, value=value)
    c.font      = Font(name="Arial", size=9, color=font_color)
    c.border    = BORDER
    c.alignment = Alignment(wrap_text=wrap, vertical="top", horizontal=align)
    if fill:
        c.fill = fill
    return c

def link_cell(ws, row, col, url, label="🔗 Open", fill=None):
    c = ws.cell(row=row, column=col, value=label)
    c.hyperlink = url
    c.font      = Font(name="Arial", size=9, color="0563C1", underline="single")
    c.border    = BORDER
    c.alignment = Alignment(horizontal="center", vertical="top")
    if fill:
        c.fill = fill
    return c

STAR_FILLS = {
    5: PatternFill("solid", fgColor="C6EFCE"),  # green
    4: PatternFill("solid", fgColor="DDEBF7"),  # blue
    3: PatternFill("solid", fgColor="FFEB9C"),  # yellow
    2: PatternFill("solid", fgColor="FFCCCC"),  # red-light
    1: PatternFill("solid", fgColor="FF9999"),  # red
}

def star_fill(stars: int):
    return STAR_FILLS.get(stars, PatternFill())

def set_col_widths(ws, widths: list[int]):
    for i, w in enumerate(widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = w


# ── Sheet 1: Top Matches ──────────────────────────────────────────────────────
def build_top_matches(ws, jobs: list[dict], today: str):
    ws.merge_cells("A1:J1")
    c = ws["A1"]
    c.value     = f"🎯 Daily Java Job Matches — Battula Bala Kotaiah | {today}"
    c.font      = Font(bold=True, color="1F3864", size=13, name="Arial")
    c.alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[1].height = 28

    ws.merge_cells("A2:J2")
    ws["A2"].value     = "Claude AI analyzed every JD against your resume  |  Stack: Java · Spring Boot · JPA · JWT · MySQL · PostgreSQL"
    ws["A2"].font      = Font(italic=True, color="555555", size=9, name="Arial")
    ws["A2"].alignment = Alignment(horizontal="center")
    ws.row_dimensions[2].height = 15

    headers = ["#", "Job Title", "Company", "Location", "Date Posted",
               "Level", "Match", "Apply Priority", "Missing Skills", "Link"]
    widths  = [4,   42,          28,         26,          14,
               16,    10,          18,              38,              12]
    set_col_widths(ws, widths)

    for ci, h in enumerate(headers, 1):
        hdr_cell(ws, 3, ci, h)
    ws.row_dimensions[3].height = 22

    for ri, job in enumerate(jobs, start=4):
        stars  = job.get("stars", 0)
        sf     = star_fill(stars)
        missing = ", ".join(job.get("skills_missing", [])) or "—"

        data_cell(ws, ri, 1,  ri - 3,                          fill=sf, align="center")
        data_cell(ws, ri, 2,  job.get("title",""),              fill=sf)
        data_cell(ws, ri, 3,  job.get("company",""),            fill=sf)
        data_cell(ws, ri, 4,  job.get("location",""),           fill=sf)
        data_cell(ws, ri, 5,  job.get("date_posted",""),        fill=sf, align="center")
        data_cell(ws, ri, 6,  job.get("seniority",""),          fill=sf)
        data_cell(ws, ri, 7,  job.get("star_label",""),         fill=sf, align="center")
        data_cell(ws, ri, 8,  job.get("apply_priority",""),     fill=sf, align="center")
        data_cell(ws, ri, 9,  missing,                          fill=sf)
        link_cell(ws, ri, 10, job.get("link",""),               fill=sf)
        ws.row_dimensions[ri].height = 32

    ws.freeze_panes = "A4"


# ── Sheet 2: Fresher Only ─────────────────────────────────────────────────────
def build_fresher_only(ws, jobs: list[dict]):
    fresher_jobs = [j for j in jobs if j.get("is_fresher_friendly")]

    ws.merge_cells("A1:H1")
    c = ws["A1"]
    c.value     = f"🟢 Fresher & Entry-Level Only — {len(fresher_jobs)} jobs found today"
    c.font      = Font(bold=True, color="375623", size=12, name="Arial")
    c.alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[1].height = 26

    headers = ["#", "Job Title", "Company", "Location", "Date Posted", "Level", "Why Fresher-Friendly", "Link"]
    widths  = [4,   42,          28,         26,          14,            16,       42,                     12]
    set_col_widths(ws, widths)

    gh = PatternFill("solid", fgColor="375623")
    for ci, h in enumerate(headers, 1):
        hdr_cell(ws, 2, ci, h, bg="375623")
    ws.row_dimensions[2].height = 22

    alt = [PatternFill("solid", fgColor="E2EFDA"), PatternFill("solid", fgColor="FFFFFF")]
    for ri, job in enumerate(fresher_jobs, start=3):
        sf = alt[(ri) % 2]
        data_cell(ws, ri, 1, ri - 2,                        fill=sf, align="center")
        data_cell(ws, ri, 2, job.get("title",""),            fill=sf)
        data_cell(ws, ri, 3, job.get("company",""),          fill=sf)
        data_cell(ws, ri, 4, job.get("location",""),         fill=sf)
        data_cell(ws, ri, 5, job.get("date_posted",""),      fill=sf, align="center")
        data_cell(ws, ri, 6, job.get("seniority",""),        fill=sf)
        data_cell(ws, ri, 7, job.get("fresher_reason",""),   fill=sf)
        link_cell(ws, ri, 8, job.get("link",""),             fill=sf)
        ws.row_dimensions[ri].height = 30

    ws.freeze_panes = "A3"


# ── Sheet 3: Skills Gap ───────────────────────────────────────────────────────
def build_skills_gap(ws, jobs: list[dict], today: str):
    ws.merge_cells("A1:D1")
    ws["A1"].value     = f"🔴 Aggregated Skills Gap — Based on {len(jobs)} jobs scraped {today}"
    ws["A1"].font      = Font(bold=True, color="C00000", size=12, name="Arial")
    ws["A1"].alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[1].height = 26

    # Count missing skills across all jobs
    missing_counter: Counter = Counter()
    for job in jobs:
        for skill in job.get("skills_missing", []):
            missing_counter[skill.strip()] += 1

    total = len(jobs)
    headers = ["Missing Skill / Technology", f"Seen in X of {total} Jobs", "Frequency %", "Priority to Learn"]
    widths  = [38, 20, 14, 45]
    set_col_widths(ws, widths)

    for ci, h in enumerate(headers, 1):
        hdr_cell(ws, 2, ci, h, bg="C00000")
    ws.row_dimensions[2].height = 22

    sorted_missing = missing_counter.most_common()

    def priority(pct):
        if pct >= 70: return "🔴 Critical — Learn ASAP"
        if pct >= 50: return "🟠 High — Learn within 1 month"
        if pct >= 30: return "🟡 Medium — Learn within 3 months"
        return "🟢 Nice to have"

    fills = {
        "🔴": PatternFill("solid", fgColor="FCE4D6"),
        "🟠": PatternFill("solid", fgColor="FFF2CC"),
        "🟡": PatternFill("solid", fgColor="FFFACD"),
        "🟢": PatternFill("solid", fgColor="E2EFDA"),
    }

    for ri, (skill, count) in enumerate(sorted_missing, start=3):
        pct     = round(count / total * 100)
        prio    = priority(pct)
        emoji   = prio[0]
        sf      = fills.get(emoji, PatternFill())

        data_cell(ws, ri, 1, skill,             fill=sf)
        data_cell(ws, ri, 2, f"{count} / {total}", fill=sf, align="center")
        data_cell(ws, ri, 3, f"{pct}%",         fill=sf, align="center")
        data_cell(ws, ri, 4, prio,              fill=sf)
        ws.row_dimensions[ri].height = 22

    ws.freeze_panes = "A3"


# ── Sheet 4: Summary ──────────────────────────────────────────────────────────
def build_summary(ws, jobs: list[dict], today: str):
    ws.column_dimensions["A"].width = 3
    ws.column_dimensions["B"].width = 65

    ws.merge_cells("A1:B1")
    ws["A1"].value     = f"📋 Daily Summary & Action Plan — {today}"
    ws["A1"].font      = Font(bold=True, color="1F3864", size=13, name="Arial")
    ws["A1"].alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[1].height = 30

    total        = len(jobs)
    fresher_cnt  = sum(1 for j in jobs if j.get("is_fresher_friendly"))
    apply_now    = [j for j in jobs if j.get("apply_priority") == "APPLY NOW"]
    strong       = [j for j in jobs if j.get("apply_priority") == "STRONG CONSIDER"]
    top5         = [j for j in jobs if j.get("stars", 0) >= 4][:5]

    all_missing: Counter = Counter()
    for job in jobs:
        for s in job.get("skills_missing", []):
            all_missing[s.strip()] += 1
    top_missing = [s for s, _ in all_missing.most_common(5)]

    lines = [
        ("TODAY'S SNAPSHOT", None, "1F3864", True),
        ("", f"Total jobs scraped & analyzed by Claude: {total}", None, False),
        ("", f"Fresher-friendly jobs: {fresher_cnt}", None, False),
        ("", f"APPLY NOW recommendations: {len(apply_now)}", None, False),
        ("", f"STRONG CONSIDER recommendations: {len(strong)}", None, False),
        ("", "", None, False),

        ("🏆 TOP 5 MATCHES TODAY", None, "375623", True),
    ]
    for i, j in enumerate(top5, 1):
        lines.append(("", f"{i}. {j.get('title','')} — {j.get('company','')} | {j.get('star_label','')} | {j.get('apply_priority','')}", None, False))

    lines += [
        ("", "", None, False),
        ("🔴 TOP SKILLS YOU ARE STILL MISSING", None, "C00000", True),
    ]
    for s in top_missing:
        lines.append(("", f"❌  {s}", None, False))

    lines += [
        ("", "", None, False),
        ("✅ ACTION PLAN (Do These Today)", None, "375623", True),
        ("", "1️⃣  Apply to every APPLY NOW job first — do it before 12 PM today.", None, False),
        ("", "2️⃣  Add Git/GitHub mentions to your resume if not already there.", None, False),
        ("", "3️⃣  Write at least one JUnit test per project and push to GitHub.", None, False),
        ("", "4️⃣  Spend 30 mins learning Docker basics (youtube: 'Docker in 100 seconds').", None, False),
        ("", "5️⃣  Check the Skills Gap sheet — focus on 🔴 Critical items this week.", None, False),
        ("", "", None, False),
        ("📧 AUTOMATION INFO", None, "1F3864", True),
        ("", "This report was auto-generated by Claude AI + Apify LinkedIn Scraper.", None, False),
        ("", "It runs every day at 10:00 AM IST via GitHub Actions.", None, False),
        ("", "Jobs are fresh — scraped from the last 24 hours only.", None, False),
    ]

    sec_fills = {
        "1F3864": PatternFill("solid", fgColor="D9E1F2"),
        "C00000": PatternFill("solid", fgColor="FCE4D6"),
        "375623": PatternFill("solid", fgColor="E2EFDA"),
    }

    row = 2
    for label, text, color, is_header in lines:
        if is_header:
            ws.merge_cells(f"A{row}:B{row}")
            c = ws.cell(row=row, column=1, value=f"  {label}")
            c.font      = Font(bold=True, color=color, size=11, name="Arial")
            c.fill      = sec_fills.get(color, PatternFill())
            c.alignment = Alignment(vertical="center")
            ws.row_dimensions[row].height = 22
        else:
            ws.cell(row=row, column=1, value="")
            c = ws.cell(row=row, column=2, value=text)
            c.font      = Font(name="Arial", size=10)
            c.alignment = Alignment(wrap_text=True, vertical="top")
            ws.row_dimensions[row].height = 18 if text else 8
        row += 1


# ── Main entry point ──────────────────────────────────────────────────────────
def generate(analyzed_jobs: list[dict], output_path: str = "daily_jobs.xlsx") -> str:
    today = date.today().strftime("%B %d, %Y")
    wb    = Workbook()

    ws1 = wb.active
    ws1.title = "Top Job Matches"
    build_top_matches(ws1, analyzed_jobs, today)

    ws2 = wb.create_sheet("Fresher Only 🟢")
    build_fresher_only(ws2, analyzed_jobs)

    ws3 = wb.create_sheet("Skills Gap 🔴")
    build_skills_gap(ws3, analyzed_jobs, today)

    ws4 = wb.create_sheet("Summary & Plan")
    build_summary(ws4, analyzed_jobs, today)

    wb.save(output_path)
    print(f"[excel] Saved → {output_path}")
    return output_path


if __name__ == "__main__":
    import json, sys
    data = json.loads(Path(sys.argv[1]).read_text()) if len(sys.argv) > 1 else []
    generate(data, "test_output.xlsx")
